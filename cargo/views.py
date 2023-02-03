#django导入部分
from django.shortcuts import render,redirect
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login,logout

#python基本库导入部分
import re
import datetime
import uuid
import requests

#第三方库库导入部分
from bs4 import BeautifulSoup
import pdfplumber
import openpyxl

#从cargo/func中导入处理相关的函数
#send_mail模块
from .func import send_mail
#从pdf中提取字符串
from .func import getstrfrompdf
#将到达wa（含板箱信息）的字符串转为相关内容
from .func import extract_wa
#从air cargo manifest中提取数据
from .func import air_cargo_manifest
#cargosalesreport pdf 和 excel对比
from .func import pdf_excel_cross_check
#cargosalesreport excel 和 excel对比
from .func import excel_excel_check


#导入models中的数据库
from .models import Awb_distribution,Awb_info

def cgo_homepage(request):
    return render(request,'cgo_homepage_templates.html')

def cgo_desk_homepage(request):
    return render(request,'cgo_desk_homepage_templates.html')

def cgo_desk_importspecialcgo_upload(request):
    return render(request,'cgo_desk_importspecialcgo_upload_templates.html')

def cgo_desk_importspecialcgo_result(request):
    if request.method=="POST":
        myfile=request.FILES.get("Weight Arrival with ULDs")
        #读取的myfile要转为二进制文件才能被pdfplumber打开
        #转为二进制文件的方法已经包含在getstrfrompdf中
        str=getstrfrompdf(myfile.read())
        flightnumber,flightdate,arr_weight,jsp_list,jca_list,temperature_list,bup_list,jph_list,special_uld_list,special_code_on_flight,trst_list,val_list,hum_list,avi_list=extract_wa(str)

        #增加
        jsp_arr_uld=[]
        for item in jsp_list:
            for uld_detail in item[2]:
                jsp_arr_uld.append(uld_detail[0])

        jsp_arr_uld_list=list(set(jsp_arr_uld))


        context={
            'flightnumber':flightnumber,
            'flightdate':flightdate,
            'arr_weight':arr_weight,
            'jsp_list':jsp_list,
            'jca_list':jca_list,
            'temperature_list':temperature_list,
            'bup_list':bup_list,
            'jph_list':jph_list,
            'special_uld_list':special_uld_list,
            'special_code_on_flight':list(set(special_code_on_flight)),
            'trst_list':trst_list,
            'val_list':val_list,
            'hum_list':hum_list,
            'avi_list':avi_list,
            'jsp_arr_uld_list':jsp_arr_uld_list,
        }
        
    return render(request,'cgo_desk_importspecialcgo_result_templates.html',context)






def cgo_fr_homepage(request):
    return render(request,'cgo_fr_homepage_templates.html')

def cgo_fr_cargosalesreport_upload(request):
    return render(request,'cgo_fr_cargosalesreport_upload_templates.html')

def cgo_fr_cargosalesreport_result(request):
    if request.method=="POST":
        file1=request.FILES.get('Cargo Sales Report 01')
        str1=getstrfrompdf(file1.read()).split("\n")
        dic1={}
        for item in str1:
            if "131-" in item:
                dic1[item[:13]]=item

        file2=request.FILES.get('Cargo Sales Report 02')
        str2=getstrfrompdf(file2.read()).split("\n")
        dic2={}
        for item in str2:
            if "131-" in item:
                dic2[item[:13]]=item

    set1=set(dic1.keys())
    set2=set(dic2.keys())

    in_set1_not_set2=set1 - set2
    in_set2_not_set1=set2 - set1
    in_both = set1 & set2

    in_both_not_equal=[]
    for item in in_both:
        if dic1[item] != dic2[item]:
            in_both_not_equal.append(item)

    context={
        'file1':file1.name,
        'file2':file2.name,
        'dic1':dic1,
        'dic2':dic2,
        'in_set1_not_set2':list(in_set1_not_set2),
        'in_set2_not_set1':list(in_set2_not_set1),
        'in_both_not_equal':in_both_not_equal,
        'empty_list':[],
    }
    return render(request,'cgo_fr_cargosalesreport_result_templates.html',context)


def cgo_traffic_homepage(request):
    return render(request,'cgo_traffic_homepage_templates.html')

def cgo_traffic_scsforotherairlines_upload(request):
    return render(request,'cgo_traffic_scsforotherairlines_upload_templates.html')

def cgo_traffic_scsforotherairlines_result(request):

    if request.method=="POST":
        america_list=[
        'BOS','ORD','JFK','YVR','PDX','SEA','YYT','YUL','YHZ','YYC',
        'YYZ','YEG','DFW','AUS','LRD','MIA','IAH','ATL','MEM','SAT','MSY','BUF',
        'BWI','CHS','CLE','CLT','CMH','IAD','MCO','ORF','PHL','RDU',
        'ROC','TPA','JAX','BNA','CVG','DTW','IND','MCI',
        'MSP','PIT','SDF','STL','LAX','SLC','SFO','DEN','ELP','LAS','PHX','SAN',
        ]
        #wd的出发内容
        file1=request.FILES.get('weightdep')
        str1=getstrfrompdf(file1.read()).split("\n")
        #获取航班信息
        flightinfo=re.search('[0-9,A-Z]{2} {0,4}[0-9]{2,5}/[0-9]{1,2}-[A-Z]{3}-20[0-9]{2}',str1[4])[0].split('/')
        flightnumber=flightinfo[0]
        flightdate=flightinfo[1]
        #获得星期几的结果，用于CK航班的号码确认
        week_day=datetime.datetime.strptime(flightinfo[1],"%d-%b-%Y").weekday()
        week_day_dic={
            0:"CK 253",
            1:"2",
            2:"CK 241",
            3:"4",
            4:"CK 241",
            5:"CK 241",
            6:"7",
        }


        if flightnumber=="JL 6744":
            flightnumber=week_day_dic[week_day]
        #获取运单信息
        # 需要发送scs的部分       
        scslist=[]
        #总运单数
        total_jlawb=0
        #总的日航重量
        total_jlweight=0
        #所有日航运单list,不包含邮件
        total_awb_list=[]
        for item in str1:         
            if item[:4]=='131-':
                total_awb_list.append(item[:12])
                total_jlawb+=1
                if item[12:15]==" P ":
                    total_jlweight+=float(item.split(' ')[3].split("/")[0])
                else:
                    total_jlweight+=float(re.match('[0-9]{3}-[0-9]{8} [0-9]{1,} [0-9]{1,}.[0-9]{1,}',item)[0].split(" ")[-1])
                dstn = re.search(' [A-Z]{3}-[A-Z]{3}',item)[0].split(" ")[-1][4:]
                

                if dstn in america_list:
                    scslist.append(item[:12])

        #邮件位置的检查
        mailawb1=request.POST.getlist('mailawb1')
        mailawb2=request.POST.getlist('mailawb2')
        mailawb3=request.POST.getlist('mailawb3')
        maildstn1=request.POST.getlist('maildstn1')
        maildstn2=request.POST.getlist('maildstn2')
        maildstn3=request.POST.getlist('maildstn3')
 
        #邮件输入内容检查
        # 运单号位数
        if len(mailawb1[0]) !=8 and len(mailawb1[0]) !=0 :
            return HttpResponse('Mail1 awb number is wrong')
        if len(mailawb2[0]) !=8 and len(mailawb2[0]) !=0 :
            return HttpResponse('Mail2 awb number is wrong')
        if len(mailawb3[0]) !=8 and len(mailawb3[0]) !=0 :
            return HttpResponse('Mail3 awb number is wrong')
        #检查8位是不是数字
        #检查目的港是不是英语3个字符

        #邮件目的港的检查
        if maildstn1[0].upper() in america_list:
            scslist.append('131-'+mailawb1[0])
        if maildstn2[0].upper() in america_list:
            scslist.append('131-'+mailawb2[0])
        if maildstn3[0].upper() in america_list:
            scslist.append('131-'+mailawb3[0])

        #邮件内容
        scs_content=''
        for item in scslist:
            scs_content+=item+", "
        content=''
        if flightnumber[:2]=="CK" or flightnumber[:2]=="MU":
            content+='TO COCC\n\n'
        if flightnumber[:2]=="5X":
            content+='TO UPS\n\n'

        
        content+=flightnumber+"/"+flightdate+" 美国方面货物保函清单\n\n\n"
        content+=scs_content+"\n\n\n"
        content+="共 "+str(len(scslist))+"票\n\n\n"
        content+="日本航空"
        #发送给cocc美国保函
        cocc_result="OK"
        upsaddress='pvgffunll@jal.com,org.pvgffkic.jali@jal.com,org.pvgffk.jali@jal.com'
        ckaddress='cocc@ceair.com,pvgffunll@jal.com,org.pvgffkic.jali@jal.com,org.pvgffk.jali@jal.com'
        try:          
            if flightnumber[:2]=="5X":
                send_address=upsaddress
            if flightnumber[:2]=="CK" or flightnumber[:2]=="MU":
                send_address=ckaddress 
            # 项目地址    
            send_mail(send_address,flightnumber+"/"+flightdate+" 美国方面货物保函",content)
            # test邮箱地址
            # send_mail('eachdayachance@hotmail.com',flightnumber+"/"+flightdate+" 美国方面货物保函",content)
        except:
            cocc_result="NOT OK"

        #发送给自己的航班信息
        content_jl=flightnumber+"/"+flightdate+'\n'
        content_jl+='航班总运单数：'+str(total_jlawb)+'票\n'
        content_jl+='航班总重量：'+str(total_jlweight)+'KG\n\n'
        for item in total_awb_list:
            content_jl+=item+'\n'
        jl_result="OK"
        try:
            # 项目地址
            send_mail('pvgffunll@jal.com,org.pvgffkic.jali@jal.com,org.pvgffk.jali@jal.com',flightnumber+"/"+flightdate+" 航班信息情况",content_jl)
            # test邮箱地址
            # send_mail('eachdayachance@hotmail.com',flightnumber+"/"+flightdate+" 航班信息情况",content_jl)
        except:
            jl_result="NOT OK"

        context={
            'cocc_result':cocc_result,
            'jl_result':jl_result,
        }
    return render(request,'cgo_traffic_scsforotherairlines_result_templates.html',context)

# desk查看板箱信息
def cgo_desk_uldstorage_result(request):
    #登陆界面
    url="https://uldmanager.champ.aero/prod/acegilogin.jsp"
    headers={
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36 Edg/93.0.961.52"
    }
    session=requests.session()
    res=session.get(url,headers=headers)
    
    #登陆完成界面
    url="https://uldmanager.champ.aero/prod/j_acegi_security_check"
    headers={
    "Host": "uldmanager.champ.aero",
    "Origin": "https://uldmanager.champ.aero",
    "Referer": "https://uldmanager.champ.aero/prod/acegilogin.jsp",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36 Edg/93.0.961.52",
    }
    data={
    "j_username": "00746148",
    "j_password": "Pvg123456",
    "submit": "Submit",
    }
    res=session.post(url,headers=headers,data=data)
    
    #数据界面
    url="https://uldmanager.champ.aero/prod/stationTotals.do?method=searchStationTotals&stCodeSearchCriteria=PVG&listType=type&lastPageNumber=0&currentPageNumber=1&recordsPerPage=15&totalNumberOfRecord=0&prevPageNumber=0&flightInd=I"
    res=session.get(url)
    res=res.content.decode()
    
    #提取数据
    bs=BeautifulSoup(res,'html.parser')
    uld = bs.find_all("tr",id="rec0op0")
    uld_inventory={}
    for item in uld:
        data = item.find_all('td')
        uld_inventory[data[0].get_text().strip()]=data[2].get_text().strip()
   
    uld_list=[]
    for key in uld_inventory:
        uld_list.append(key.ljust(10)+uld_inventory[key].rjust(4))
        uld_list.append(' '.rjust(14))

    html_dict={
        "uld_list":uld_list,
    }
    return render(request,"cgo_desk_uldstorage_result_templates.html",html_dict)

#ic homepage
def cgo_ic_homepage(request):
    return render(request,'cgo_ic_homepage_templates.html')

#ULD cross check 逻辑
def cgo_ic_crosscheck_upload(request):
    return render(request,'cgo_ic_crosscheck_upload_templates.html',{})

def cgo_ic_crosscheck_result(request):
    if request.method == 'POST':
        li=request.FILES.get('loadinginstruction')
        ex=request.FILES.get('warehouseuldlist')

    #处理pdf
    pdfpage=pdfplumber.open(li)
    pdfstring=''
    for page in pdfpage.pages:
        pdfstring+=page.extract_text()
    type_regx=re.compile('JA[0-9]{3,4}J')#机型

    uld_regx=re.compile("[A-Z]{3}[0-9]{3,5}[A-Z]{2,3}[\s]*[A-Z]*[\s]*[0-9]*/[[\s]*[0-9]*")
    uld_list=uld_regx.findall(pdfstring)
    uld_weight_list={}
    for item in uld_list:
        uld_number_regx=re.compile("[A-Z]{3}[0-9]{3,5}[A-Z]{2,3}")
        uld=uld_number_regx.findall(item)[0]
        #修正PLA缺少一位0的错误
        if "PLA" in uld:
            uld=uld.replace("PLA","PLA0")
        ###########
        uld_weight_regx=re.compile("/[[\s]*[0-9]*")
        weight=int(uld_weight_regx.findall(item)[0][1:])
        uld_weight_list[uld]=weight
    pdfpage.close()

    wb=openpyxl.load_workbook(ex,data_only=True)
    ws = wb['交接单']
    uld_list_left = ws['c5:d21']
    uld_list_right = ws['j5:k21']
    uld_check_list_set = {}
    for row in uld_list_left:
        if row[0].value != None:
            uld_check_list_set[row[0].value] = row[1].value
    for row in uld_list_right:
        if row[0].value != None:
            uld_check_list_set[row[0].value] = row[1].value

    return_str=[]
    for key in uld_weight_list:
        if key in uld_check_list_set:
            if uld_weight_list[key] != uld_check_list_set[key]:
                return_str.append(key+' weight may be wrong.')
        else:
            return_str.append(key+' uld number may be wrong or in BLK')
    if return_str==[]:
        return_str.append('all check ok')

    context={'return_str':return_str}

    return render(request,'cgo_ic_crosscheck_result_templates.html',context)



def cgo_traffic_notallowedcargo_upload(request):
    return render(request,'cgo_traffic_notallowedcargo_upload_templates.html')

#尚未完成
def cgo_traffic_notallowedcargo_result(request):
    if request.method == 'POST':
        manifest=request.FILES.get('manifest')
    manifest_details=air_cargo_manifest(manifest)
    cargo_not_allowed_list=[]
    for key in manifest_details:
        #处理有品名检查的机制
        if manifest_details[key][2]==True:
           cargo_not_allowed_list.append("MAWB: "+key)
        if manifest_details[key][4]:
            for hawb in manifest_details[key][4]:
                if manifest_details[key][4][hawb][0]==True:
                    cargo_not_allowed_list.append("MAWB: "+key+"     HAWB: "+hawb)

    context={
        'cargo_not_allowed_list':cargo_not_allowed_list,
    }    
  
    return render(request,'cgo_traffic_notallowedcargo_result_templates.html',context)


def cgo_fr_cargosalesreport_pdfexcel_upload(request):
    return render(request,"cgo_fr_cargosalesreport_pdfexcel_upload_templates.html")

def cgo_fr_cargosalesreport_pdfexcel_result(request):
    if request.method=="POST":
        pdf=request.FILES.get('Cargo Sales Report pdf')
        xlsx=request.FILES.get('Cargo Sales Report excel')
    inpdfnotexcel,inexcelnotpdf,notequal=pdf_excel_cross_check(pdf,xlsx)
    context={
        "inpdfnotexcel":inpdfnotexcel,
        "inexcelnotpdf":inexcelnotpdf,
        "notequal":notequal,
    }  
    return render(request,"cgo_fr_cargosalesreport_pdfexcel_result_templates.html",context)

def cgo_fr_cargosalesreport_excelexcel_upload(request):
    return render(request,'cgo_fr_cargosalesreport_excelexcel_upload_templates.html')

def cgo_fr_cargosalesreport_excelexcel_result(request):
    if request.method=="POST":
        xlsx1=request.FILES.get('xlsx1')
        xlsx2=request.FILES.get('xlsx2')
    in1not2,in2not1,not_equal_list = excel_excel_check(xlsx1,xlsx2)
    context={
        "in1not2":in1not2,
        "in2not1":in2not1,
        "not_equal_list":not_equal_list,
        "xlsx1":xlsx1.name,
        "xlsx2":xlsx2.name,
    }
    return render(request,"cgo_fr_cargosalesreport_excelexcel_result_templates.html",context)

#登录cargo页面
def cgo_login(request):
    if request.method=="POST":
        username=request.POST.get("username")
        password=request.POST.get("password")
        user = authenticate(username=username,password=password)
        if user:
            login(request,user)
            return redirect("cgo_homepage")
    return render(request,'cgo_login_templates.html')

def cgo_fr_awbdistribution_upload(request):
    return render(request,"cgo_fr_awbdistribution_upload_templates.html")

def cgo_fr_awbdistribution_result(request):
    distribution_uuid=str(uuid.uuid4())
    if request.method=="POST":
        agent=request.POST.get("agent").upper()
        awbnumber=request.POST.get("awbnumber")
        piece=int(request.POST.get("piece"))

    #awb_list 运算逻辑    
    awb_list=[]
    awb_list.append("131-"+awbnumber)
    head=int(awbnumber)
    for i in range(piece-1):

        if head % 10<6:
            head=head+11
        else:
            head=head+4
        tail=len(str(head))
        awb_list.append("131-"+"0"*(8-tail)+str(head))

    #awb_list 加入数据库
    for awb in awb_list:
        awb_info=Awb_info(number=awb,agent=agent,distribution_uuid=distribution_uuid)
        awb_info.save()

    awbdistribution_create=Awb_distribution(agent=agent,
                                            piece=piece,
                                            distribution_uuid= distribution_uuid)
    awbdistribution_create.save()
    distribution_create_date=str(awbdistribution_create.distribution_date)
 
    #等待添加运单在展示画面中的展示方式

    context={
        "agent":agent,
        "piece":piece,
        "awb_list":awb_list,
        "distribution_create_date":distribution_create_date,
        "distribution_uuid":distribution_uuid,
    }
    return render(request,"cgo_fr_awbdistribution_result_templates.html",context)

