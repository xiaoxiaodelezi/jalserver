import io
import pdfplumber
import re
import os
import uuid

#定义一些基本函数

#邮件发送模块
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

def send_mail(toAddress, subject, mail_content, file_path='', filename=[]):
    msg = MIMEMultipart()
    msg['From'] = 'jlpvgdocument@hotmail.com'
    msg['To'] = toAddress  # txt文件中的邮箱地址 添加收件人
    msg['Subject'] = subject # 添加邮件主题
    if mail_content != "":
        mail_body = mail_content  # 添加邮件内容
        msg.attach(MIMEText(mail_body, 'plain', 'utf-8'))  # 将邮件内容附加到邮件编码中
    if file_path != "" or filename !=[]:
        for file in filename:
            message = MIMEText(open(file_path + file, "rb").read(), 'base64', 'utf-8')  # 附件文件
            message.add_header('Content-Disposition', 'attachment', filename=file)  # 附件显示名称
            msg.attach(message)  # 将附件附加到邮件编码中
    # 邮件端口处理
    s = smtplib.SMTP("smtp-mail.outlook.com", 587)
    s.ehlo()  # Hostname to send for this command defaults to the fully qualified domain name of the local host.
    s.starttls()  # Puts connection to SMTP server in TLS mode
    # 发件人
    fromAddress = 'jlpvgdocument@hotmail.com'  # sender
    # 发件邮箱地址和密码
    smtpLogin = 'jlpvgdocument@hotmail.com'
    smtpPasswd = 'PVGFFU-003'
    s.login(smtpLogin, smtpPasswd)
    s.sendmail(fromAddress, toAddress.split(','), msg.as_string())
    s.quit()


#在file_pool中创建一个临时文件夹
#返回包含路径（包含最后的"/"）
def mk_tempdir():
    new_dir=str(uuid.uuid4())
    os.mkdir("./file_pool/"+new_dir)
    return "./file_pool/"+new_dir+"/"


#输入辅材信息，在file_pool的临时文件夹中生成一个辅材excel实例
#返回生成为文件名，用来加入list，统一添加为邮件附件
import openpyxl
def auxiliary_instance(wood,bandage,skid,flight,date,agency,new_dir):
    ttl_weight= bandage * 1.5 + wood * 1.6 + skid *25 
    wb=openpyxl.load_workbook('./file_templates/auxiliary_materials.xlsx')
    ws=wb["Sheet1"]
    ws['D19']=bandage
    ws['D21']=wood
    ws['D20']=skid
    ws['f22']=str(ttl_weight)+" KG"
    ws['C14']=flight+"/"+date
    ws['C16']=agency.upper()
    file_name=agency.upper()+"_"+flight+"_"+date
    wb.save("../file_pool/"+new_dir+file_name)
    wb.close()
    return file_name



# 返回一个从二进制pdf中提取的字符串
def getstrfrompdf(file):
    f=io.BytesIO(file) 
    pdf = pdfplumber.open(f)
    str_result=""
    for page in pdf.pages:
       str_result+=page.extract_text()+'\n'
    return str_result


# 进口带板箱信息的wa解析
def extract_wa(str_input):
    #以行为单位分割str
    original_list=str_input.split('\n')    
    length=(len(original_list))
    for i in range(0,length):
        #标记运单号前的锚点，方便按运单切割
        if re.match("[0-9]{3}-[0-9]{8}",original_list[i]) != None:
            original_list[i]=";"+original_list[i]

    #获取航班号，航班日期，航班到达总重量
    #for函数在str的前15行寻找正则匹配
    for i in range(15):
        if re.search('[0-9,A-Z]{2}[ ]{0,2}[0-9]{1,4}/[0-9]{1,2}-[A-Z]{3}-[0-9]{4}',original_list[i]) !=None:
                dateinfo=re.search('[0-9,A-Z]{2}[ ]{0,2}[0-9]{1,4}/[0-9]{1,2}-[A-Z]{3}-[0-9]{4}',original_list[i])[0].split('/')
                #航班号
                flightnumber=dateinfo[0]
                #航班日期
                flightdate=dateinfo[1]   
    #到达货重，整数位四舍五入
    arr_weight=round(float(original_list[-2].split(" ")[-1][:-1]))
    
    #删除不需要的信息
    #需要删除的部分必定包含delete_list中的一个字段
    delete_list=[
        "Date :",
        "Import Check Manifest",
        "Page : ",
        "Prepared by :",
        "Flight No.",
        "NATURE",
        "AIR WAYBILL PIECES",
        "OF GOODS",
        "-NIL-",
    ]
    for i in original_list[::-1]:
        for j in delete_list:
            if j in i:
                original_list.remove(i)
                break

    #标记各个进口各种运单属性的锚点，初始值为0                
    revlocal=0
    revtrst=0
    nrevlocal=0
    nrevtrst=0
    grandtotal=0

    #获取各个锚定位置的信息
    for i in range(0,len(original_list)):
        if "<REV Local>" in original_list[i]:
            revlocal=i
        if "<REV TRST>" in original_list[i]:
            revtrst=i
        if "<NREV Local>" in original_list[i]:
            nrevlocal=i
        if "<NREV TRST>" in original_list[i]:
            nrevtrst=i
        if "GRAND TOTAL" in original_list[i]:
            grandtotal=i


    #初始化各个类型运单的字符串     
    rev_local = ""
    rev_trst = ""
    nrev_local = ""
    nrev_trst= ""
    #将各个类型的运单的各行分别串起来，形成各自的字符串
    for i in range(revlocal+1,revtrst):
        rev_local = rev_local + original_list[i].strip() +" "
    rev_local=rev_local.strip()
    
    for i in range(revtrst+1,nrevlocal):
        rev_trst = rev_trst + original_list[i].strip() + " "
    rev_trst = rev_trst.strip()

    for i in range(nrevlocal+1,nrevtrst):
        nrev_local = nrev_local + original_list[i].strip() + " "
    nrev_local = nrev_local.strip()

    for i in range(nrevtrst+1,grandtotal-1):
        nrev_trst = nrev_trst + original_list[i].strip() + " "
    nrev_trst = nrev_trst.strip()


    #收集转运航班信息
    trst_list=[]
    trststr=rev_trst.strip()+" "+nrev_trst.strip()
    if trststr != "":
        trststr_list=trststr.split(";")[1:]
        for item in trststr_list:
            trstawb=item[:13]
            trstdst=re.search("[A-Z]{3}-[A-Z]{3}",item)[0].split("-")[1]
            contruck=re.search("Connection: [A-Z]{2}[ ]{0,1}[0-9]{2,4}/[0-9]{1,2}[A-Z]{3}",item)[0]
            trst_list.append([trstawb,trstdst,contruck])



    #定义一个函数，用于从一个字符串中获得单个运单的信息
    #运单号，出发到达港，特殊代码，所在板箱和重量信息
    def getdetail(str):
        # 作为返回值的货物字典
        special_dic={}
        #去掉第一个空的item
        str_list=str.split(";")[1:]
        for item in str_list:
            #找出出发港-目的港这个标记
            #item逆向找出最后一个dep-dtn的匹配
            pos=re.search(' [A-Z]{3}-[A-Z]{3} ',item[::-1]).span()
            pos=(len(item)-pos[1],len(item)-pos[0])                        
            #获取运单号
            awb=item[:12]
            #获取出发港
            deps=item[pos[0]+1:pos[0]+4]
            #获取到达港
            arrs=item[pos[0]+5:pos[0]+8]
            #sstr包含从特殊代码到板箱信息的所有部分
            sstr=item[pos[1]:].strip()

            #获取运单shc代码个数
            sp_index=0
            while True:
                #循环直至第n*4的位置不是空格，代表特殊代码结束
                if sstr[sp_index*4+3]==" ":
                    sp_index=sp_index+1
                else:
                    break
            
            #初始化特殊代码字段
            shc_str=''
            #初始化板箱信息字段
            uld_str=''
            #如果特殊代码个数为0，那么所有sstr都是板箱字段
            if sp_index == 0:
                uld_str=sstr
            #其余情况是sstr去掉特殊代码部分余下的字段就是板箱字段
            else:
                uld_str=sstr[sp_index*4:]
                shc_str=sstr[:sp_index*4-1]
            
            #获得shc的各个代码，以列表形式返回
            shc_code=shc_str.split(' ')
            #板箱字符串通过空格分隔成一个列表
            uld_l=uld_str.split(' ')
            #列表中3个一组，第一个为板箱号，第二个是件数，第三个是重量
            uld_n=int(len(uld_l)/3)
            #这票运单的板箱信息，key：板箱号，value是(这个板箱上的件数，这个板箱上的重量)
            uld_dic={}
            #初始化这票运单本航班的到达总件数
            arr_total_p=0
            #初始化这票运单本航班的到达总重量
            arr_total_weight=0

            #循环板箱列表，获取信息
            for i in range(uld_n):
                uld_dic[uld_l[i*3]]=(uld_l[i*3+1],uld_l[i*3+2])
                #统计这票运单的到达总件数
                arr_total_p+=int(uld_l[i*3+1])
                #统计这票运单的到达总重量
                arr_total_weight+=float(uld_l[i*3+2])
            #round到小数点一位，防止浮点运算造成的误差
            arr_total_weight=round(arr_total_weight,1)

            #构建返回信息
            special_dic[awb]=[
                #list形式，包含special
                shc_code,
                #dict形式，板箱号:(件数，重量)
                uld_dic,
                #出发港
                deps,
                #目的港
                arrs,
                #本次航班到达总件数
                arr_total_p,
                #本次航班到达总重
                arr_total_weight,                
            ]
        return special_dic

    # 总共返回以下几个信息
    # 1.航班号    str
    # 2.航班日期  str
    # 3.到达重量  float
    # 4.rev_local的运单信息   dic[awb]=list
    # 5.rev_trst的运单信息    dic[awb]=list
    # 6.nrev_local的运单信息  dic[awb]=list
    # 7.nrev_trst的运单信息   dic[awb]=list

    jsp_list=[]
    jca_list=[]
    temperature_list=[]
    bup_list=[]
    jph_list=[]
    special_uld_list=[]
    special_code_on_flight=[]

    avi_list=[]
    hum_list=[]
    val_list=[]    
    special_uld_set=['RAP',"AKN","RKN"]
    temperature_request_list=['COL',"PER","ICE","JPH"]



    for awb_kind in [rev_local,nrev_local]:
        awb_kind_dic=getdetail(awb_kind)

        for key in getdetail(awb_kind):
            for sp in awb_kind_dic[key][0]:
                special_code_on_flight.append(sp)

            for suld in special_uld_set:
                for item in awb_kind_dic[key][1]:
                    if suld in item:
                        special_uld_list.append(item)

            if "JSP" in awb_kind_dic[key][0]:
                awb_uld_list=[]
                for uld in awb_kind_dic[key][1]:
                    for i in special_uld_set:
                        if i in uld:
                            special_uld_list.append(uld)
                    awb_uld_list.append([uld,awb_kind_dic[key][1][uld][0],awb_kind_dic[key][1][uld][1]])
                jsp_list.append([key,awb_kind_dic[key][0],awb_uld_list,awb_kind_dic[key][2],awb_kind_dic[key][3],awb_kind_dic[key][4],awb_kind_dic[key][5]])
            
            if "BUP" in awb_kind_dic[key][0]:
                awb_uld_list=[]
                for uld in awb_kind_dic[key][1]:
                    for i in special_uld_set:
                        if i in uld:
                            special_uld_list.append(uld)
                    awb_uld_list.append([uld,awb_kind_dic[key][1][uld][0],awb_kind_dic[key][1][uld][1]])
                bup_list.append([key,awb_kind_dic[key][0],awb_uld_list,awb_kind_dic[key][2],awb_kind_dic[key][3],awb_kind_dic[key][4],awb_kind_dic[key][5]])
        
            if "JCA" in awb_kind_dic[key][0]:
                awb_uld_list=[]
                for uld in awb_kind_dic[key][1]:
                    for i in special_uld_set:
                        if i in uld:
                            special_uld_list.append(uld)
                    awb_uld_list.append([uld,awb_kind_dic[key][1][uld][0],awb_kind_dic[key][1][uld][1]])
                jca_list.append([key,awb_kind_dic[key][0],awb_uld_list,awb_kind_dic[key][2],awb_kind_dic[key][3],awb_kind_dic[key][4],awb_kind_dic[key][5]])

            if "JPH" in awb_kind_dic[key][0]:
                awb_uld_list=[]
                for uld in awb_kind_dic[key][1]:
                    for i in special_uld_set:
                        if i in uld:
                            special_uld_list.append(uld)
                    awb_uld_list.append([uld,awb_kind_dic[key][1][uld][0],awb_kind_dic[key][1][uld][1]])
                jph_list.append([key,awb_kind_dic[key][0],awb_uld_list,awb_kind_dic[key][2],awb_kind_dic[key][3],awb_kind_dic[key][4],awb_kind_dic[key][5]])
    
            if "HUM" in awb_kind_dic[key][0]:
                awb_uld_list=[]
                for uld in awb_kind_dic[key][1]:
                    for i in special_uld_set:
                        if i in uld:
                            special_uld_list.append(uld)
                    awb_uld_list.append([uld,awb_kind_dic[key][1][uld][0],awb_kind_dic[key][1][uld][1]])
                hum_list.append([key,awb_kind_dic[key][0],awb_uld_list,awb_kind_dic[key][2],awb_kind_dic[key][3],awb_kind_dic[key][4],awb_kind_dic[key][5]])
    
            if "VAL" in awb_kind_dic[key][0]:
                awb_uld_list=[]
                for uld in awb_kind_dic[key][1]:
                    for i in special_uld_set:
                        if i in uld:
                            special_uld_list.append(uld)
                    awb_uld_list.append([uld,awb_kind_dic[key][1][uld][0],awb_kind_dic[key][1][uld][1]])
                val_list.append([key,awb_kind_dic[key][0],awb_uld_list,awb_kind_dic[key][2],awb_kind_dic[key][3],awb_kind_dic[key][4],awb_kind_dic[key][5]])
    
            if "AVI" in awb_kind_dic[key][0]:
                awb_uld_list=[]
                for uld in awb_kind_dic[key][1]:
                    for i in special_uld_set:
                        if i in uld:
                            special_uld_list.append(uld)
                    awb_uld_list.append([uld,awb_kind_dic[key][1][uld][0],awb_kind_dic[key][1][uld][1]])
                avi_list.append([key,awb_kind_dic[key][0],awb_uld_list,awb_kind_dic[key][2],awb_kind_dic[key][3],awb_kind_dic[key][4],awb_kind_dic[key][5]])
    
            for temperature_request in temperature_request_list :   
                if temperature_request in awb_kind_dic[key][0]:
                    awb_uld_list=[]
                    for uld in awb_kind_dic[key][1]:
                        for i in special_uld_set:
                            if i in uld:
                                special_uld_list.append(uld)
                        awb_uld_list.append([uld,awb_kind_dic[key][1][uld][0],awb_kind_dic[key][1][uld][1]])
                    temperature_list.append([key,awb_kind_dic[key][0],awb_uld_list,awb_kind_dic[key][2],awb_kind_dic[key][3],awb_kind_dic[key][4],awb_kind_dic[key][5]])
                    break      

    #修改了special的逻辑，考虑返回值的使用
    return flightnumber,flightdate,arr_weight,jsp_list,jca_list,temperature_list,bup_list,jph_list,list(set(special_uld_list)),special_code_on_flight,trst_list,val_list,hum_list,avi_list


def air_cargo_manifest(manifest):

    not_allowed_list=['SCOPE','SEA',"POWERPACK"]

    pdf = pdfplumber.open(manifest)
    pdf_list=[]
    for page in pdf.pages:
        for page_str in page.extract_text().split('\n')[9:]:
            pdf_list.append(page_str)
    pdf_list=pdf_list[:-2]
    total_line=len(pdf_list)
    awb_mark=[]
    for i in range(len(pdf_list)):
        if re.match('[0-9]{3}-[0-9]{8} ',pdf_list[i]):
            awb_mark.append(i)
    awb_mark.append(total_line)

    # dic[awb]=第二行的开始7个字符，第四行国家内容，dic[houseawb]/(限制品(torf),最后一行国家内容)
    manifest_detail={}

    total_awbmark=len(awb_mark)
    for i in range(total_awbmark-1):
        awbnumber=pdf_list[awb_mark[i]][:12]
        awb_dep=pdf_list[awb_mark[i]+1][:3]
        awb_arr=pdf_list[awb_mark[i]+1][4:7]
        awb_country=pdf_list[awb_mark[i]+3]
        awb_not_allowed_result=False
        for item in not_allowed_list:
            if item in pdf_list[awb_mark[i]]:
                awb_not_allowed_result=True
                break

        if awbnumber in manifest_detail:
            hawb_dic=manifest_detail[awbnumber][4]

        hawb_dic={}
        for k in range(1,int((awb_mark[i+1]-awb_mark[i])/4)):
            hawb_number = pdf_list[awb_mark[i]+4*k].split(" ")[0]
            not_allowed_result=False
            if hawb_number not in hawb_dic:
                for item in not_allowed_list:
                    if item in pdf_list[awb_mark[i]+4*k]:
                        not_allowed_result=True
                        break
                hawb_country=pdf_list[awb_mark[i]+4*k+3]
                hawb_dic[hawb_number]=[not_allowed_result,hawb_country]
        manifest_detail[awbnumber]=[
            awb_dep,
            awb_arr,
            awb_not_allowed_result,
            awb_country,
            hawb_dic
        ]
    
    return manifest_detail


import decimal
#pdf 和 excel的cross check
def pdf_excel_cross_check(pdf,xlsx):            
    pdf=pdfplumber.open(pdf)
    pdf_awb_dic={}
    for page in pdf.pages:
        for line in page.extract_text().split("\n"):
            if line[:4] == "131-":
                if "/" in line:
                    pdf_awb_dic[line[:3]+line[4:12]]=" ".join(line[13:].split(' ')[:-1]).replace(",","")
                else:
                    pdf_awb_dic[line[:3]+line[4:12]]=" ".join(line[13:].split(' ')).replace(",","")

    wb=openpyxl.load_workbook(xlsx,data_only=True)
    ws=wb.worksheets[0]
    excel_awb_dic={}
    for i in range (2,ws.max_row+1):
        # awbnum=ws.cell(i,13).value
        v=decimal.Decimal(ws.cell(i,22).value).quantize(decimal.Decimal('0.00'))
        ac=decimal.Decimal(ws.cell(i,29).value).quantize(decimal.Decimal('0.00'))
        y=decimal.Decimal(ws.cell(i,25).value).quantize(decimal.Decimal('0.00'))
        v_ab=decimal.Decimal(float(ws.cell(i,22).value)-float(ws.cell(i,28).value)).quantize(decimal.Decimal('0.00'))
        ab__ac=decimal.Decimal(float(ws.cell(i,28).value)+float(ws.cell(i,29).value)).quantize(decimal.Decimal('0.00'))#ab+ac
        i_j=ws.cell(i,9).value+"-"+ws.cell(i,10).value#i-j
        x=decimal.Decimal(float(ws.cell(i,24).value)).quantize(decimal.Decimal('0.00'))#x
        excel_awb_str=str(v)+" "+str(ac)+" "
        if y:
            excel_awb_str+=str(y)+" "
        if v_ab:
            excel_awb_str+=str(v_ab)+" "
        excel_awb_str+= str(ab__ac)+" "+i_j+" "+str(x)
        excel_awb_dic[ws.cell(i,13).value]=excel_awb_str

    setpdf=set(pdf_awb_dic.keys())
    setexcel=set(excel_awb_dic.keys())
    inpdfnotexcel=list(setpdf-setexcel)
    inexcelnotpdf=list(setexcel-setpdf)
    notequal=[]
    for item in list(setexcel & setpdf):
        if excel_awb_dic[item] != pdf_awb_dic[item]:
            notequal.append(item)
    return inpdfnotexcel,inexcelnotpdf,notequal

def excel_excel_check(xlsx1,xlsx2):     
    check_col_list=[22,24,25,26,28,29]       
    wb1=openpyxl.load_workbook(xlsx1,data_only=True)
    ws=wb1.worksheets[0]
    xlsx1_awb_dic={}
    for i in range (2,ws.max_row+1):
        add_list=[]
        for col in check_col_list:
            add_list.append(ws.cell(i,col).value)
        xlsx1_awb_dic[ws.cell(i,13).value]=add_list
    wb1.close()

    wb2=openpyxl.load_workbook(xlsx2,data_only=True)
    ws=wb2.worksheets[0]
    xlsx2_awb_dic={}
    for i in range (2,ws.max_row+1):
        add_list=[]
        for col in check_col_list:
            add_list.append(ws.cell(i,col).value)
        xlsx2_awb_dic[ws.cell(i,13).value]=add_list
    wb2.close()

    xlsx1_set=set(xlsx1_awb_dic.keys())
    xlsx2_set=set(xlsx2_awb_dic.keys())
    in1not2=list(xlsx1_set-xlsx2_set)
    in2not1=list(xlsx2_set-xlsx1_set)
    not_equal_list=[]

    for awb in (xlsx1_set & xlsx2_set):
        if xlsx1_awb_dic[awb] != xlsx2_awb_dic[awb]:
            not_equal_list.append(awb)
    return in1not2,in2not1,not_equal_list